import os
import csv
import time
import random
import requests
import socket
import socks
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver import Chrome
from docx import Document
from docx.shared import Inches
from colorama import init as colorinit
from colorama import Fore, Back, Style
from openpyxl import Workbook
from openpyxl import load_workbook

colorinit()

browser = False

def create_folder(path):
	try:
		os.makedirs(path)
	except:
		print(Style.BRIGHT, Fore.YELLOW, 
			'\nПапка существует\t=====>\t [ WARNING ]', Style.RESET_ALL)

def delay():
	sec = random.uniform(2.0, 3.0)
	print('Delay: ' + str(sec) + ' sec\n')
	time.sleep(sec)

def get_id(url):
	global adID
	adID = url.split('_')[-1]

def launch_browser():
	global driver

	print(Style.BRIGHT, Fore.CYAN + '\nИнициализация браузера...' + Style.RESET_ALL)
	chromeOptions = webdriver.ChromeOptions()
	chromeOptions.add_argument('headless')
	chromeOptions.add_argument('--ignore-certificate-errors')
	chromeOptions.add_argument('--test-type')
	chromeOptions.add_argument('--log-level=3')
	chromeOptions.add_argument('--window-size=1000x3000')
	driver = webdriver.Chrome(options=chromeOptions)

	print(Style.BRIGHT, Fore.GREEN, '\n Браузер запущен\t=====>\t [ OK ]' + Style.RESET_ALL)
	return driver

def open_page(url):
	global driver

	try:
		driver.get(url)
		print(Style.BRIGHT, Fore.GREEN + 'Страница открыта\t=====>\t [ OK ]' + Style.RESET_ALL)
	except:
		print(Style.BRIGHT, Fore.RED + 'Страница не открыта\t=====>\t [ ERROR ]' + Style.RESET_ALL)
	
	try:
		driver.find_element_by_class_name('item-phone-number').click()
		time.sleep(2) # !!!

		driver.find_elements_by_class_name('close')[-1].click()
		print(Style.BRIGHT, Fore.GREEN + 'Номер телефона открыт\t=====>\t [ OK ]' + Style.RESET_ALL)
	except:
		print(Style.BRIGHT, Fore.RED + 'Номер тел. не открыт\t=====>\t [ ERROR ]' + Style.RESET_ALL)

	try:
		driver.find_element_by_class_name('item-map-control').click()
		print(Style.BRIGHT, Fore.GREEN + 'Карта скрыта\t\t=====>\t [ OK ]' + Style.RESET_ALL)
	except:
		print(Style.BRIGHT, Fore.RED + 'Карта не скрыта\t=====>\t [ ERROR ]' + Style.RESET_ALL)
	
def take_screenshot():
	global driver

	driver.save_screenshot('img.png')
	print(Style.BRIGHT, Fore.GREEN + 'Скриншот сохранен\t=====>\t [ OK ]' + Style.RESET_ALL)

def create_document():
	global adID
	global path

	doc = Document()
	doc.add_picture('img.png', width=Inches(4))
	doc.save(path + adID + '.docx')
	print(Style.BRIGHT, Fore.GREEN + 'Документ сохранен\t=====>\t [ OK ]' + Style.RESET_ALL)

def write_xlsx(data):
	global filename
	global path

	try:
		file = open(path + filename + '.xlsx')
	except IOError as e:
		wb = Workbook()
		ws = wb.active
		wb.save(path + filename + '.xlsx')

	book = load_workbook(path + filename + '.xlsx')
	sheet = book.active

	row = ( (
		data['title'],
		data['location'],
		data['square'],
		data['price'],
		data['adID'] ) )

	sheet.append(row)

	book.save(path + filename + '.xlsx')
	print(Style.BRIGHT, Fore.GREEN + 'Записано в .xlsx\t=====>\t [ OK ]' + Style.RESET_ALL)

'''
def write_csv(data):
	with open('avito.csv', 'a', encoding='utf8') as f:
		writer = csv.writer(f)

		writer.writerow( (data['title'],
						  data['location'],
						  data['square'],
						  data['price'],
						  data['adID']) )
'''

def get_total_pages(url):
	try:
		html = requests.get(url, headers={'User-Agent': UserAgent().chrome}).text
		soup = BeautifulSoup(html, 'lxml')
		pages = soup.find('div', class_='pagination-pages').find_all('a', class_='pagination-page')[-1].get('href')
		total_pages = pages.split('=')[1].split('&')[0]
	except:
		total_pages = 1

	return int(total_pages)

def logic_hub(url, page, adnum):
	global browser

	html = requests.get(url, headers={'User-Agent': UserAgent().chrome}).text
	soup = BeautifulSoup(html, 'lxml')
	ads = soup.find('div', class_='js-catalog-list').find_all('div', class_='item_table')

	for ad in ads:
		url = 'https://www.avito.ru' + ad.find('div', class_='item_table-header').find('h3').find('a').get('href')
		
		get_id(url)

		if browser == False:
			launch_browser()
			print(Style.BRIGHT, 
				  Fore.WHITE,
				  '\n' + str(page) + ':' + str(adnum) + ' - ', Fore.MAGENTA + url, 
				  Style.RESET_ALL)
			open_page(url)
			take_screenshot()
			create_document()
			browser = True
		else:
			print(Style.BRIGHT, 
				  Fore.WHITE,
				  '\n' + str(page) + ':' + str(adnum) + ' - ', Fore.MAGENTA + url, 
				  Style.RESET_ALL)
			open_page(url)
			take_screenshot()
			create_document()

		html = requests.get(url, headers={'User-Agent': UserAgent().chrome}).text
		soup = BeautifulSoup(html, 'lxml')

		# Title
		try:
			title = soup.find('div', class_='title-info-main').text.strip()
		except:
			title = ''
			print(Style.BRIGHT, Fore.YELLOW + 'Заголовок отсутствует\t=====>\t [ WARNING ]' + Style.RESET_ALL)

		# Location
		try:
			location = soup.find('div', class_='item-map-location').text.strip().replace('\n', '').replace('  ', ' ').split('Скрыть карту')[0].split('Адрес: ')[-1]
		except:
			try:
				location = soup.find('div', class_='item-view-seller-info').text.split('Адрес')[-1].strip()
			except:
				location = ''
				print(Style.BRIGHT, Fore.YELLOW + 'Местоположение отсутствует\t=====>\t [ WARNING ]' + Style.RESET_ALL)

		# Square
		try:
			square = soup.find('div', class_='item-params').text.strip().replace('\n', '').replace('\xa0', '').split('м²')[0].split(';')[0].split(':')[-1].split(' ')[-1].replace('кирпичный', '').replace('металлический', '').replace('железобетонный', '')
		except:
			square = ''
			print(Style.BRIGHT, Fore.YELLOW + 'Площадь отсутствует\t=====>\t [ WARNING ]' + Style.RESET_ALL)

		# Price
		try:
			price = soup.find('div', class_='item-price').find('span', class_='price-value-string').text.strip().replace('\xa0', '').split('₽')[0].replace('Не указана', '').replace('Договорная', '')
		except:
			price = ''
			print(Style.BRIGHT, Fore.YELLOW + 'Цена отсутствует\t=====>\t [ WARNING ]' + Style.RESET_ALL)

		# ID
		try:
			adID = soup.find('div', class_='title-info-metadata').text.split()[1].split(',')[0]
		except:
			adID = ''
			print(Style.BRIGHT, Fore.YELLOW + 'ID отсутствует\t=====>\t [ WARNING ]' + Style.RESET_ALL)

		data = {'title': title,
				'location': location,
				'square': square,
				'price': price,
				'adID': adID}

		write_xlsx(data)
		#write_csv(data)
		adnum += 1

		#delay() !!!!!!!!!!

def main():
	global filename
	global path

	print('\n=============================================',
		'\nStarlight United v1.5 (14/12/2018)',
		'\nРазработчик: Чернышев Егор Владимирович',
		'\n\nФункционал ограничен',
		'\n=============================================')

	filename = input('\nВведите имя Папки/Файла: ').replace(' ', '_').replace(',', '_')
	if filename == '':
		filename = 'Starlight_United_Untitled'
		print(Style.BRIGHT + Fore.YELLOW + '\nИмя не указано. Файлы будут сохранены в папку',
			'"' + filename + '"', Style.RESET_ALL)
	path = filename + '\\'

	create_folder(path)

	p_url = input('\nВведите ссылку: ')
	f_url = p_url + '&p='

	try:
		startup_page = int(input('\nВведите номер страницы, с которой хотите начать: '))
	except:
		startup_page = 1
		print('\nНачальная страница не указана. Выбрана страница по умолчанию',
			Style.BRIGHT + Fore.YELLOW + '(1)')

	total_pages = get_total_pages(p_url)

	print(Style.BRIGHT + Fore.GREEN + '\nПРОГРАММА НАЧАЛА ВЫПОЛНЕНИЕ РАБОТЫ' + Style.RESET_ALL)

	for i in range(startup_page, total_pages + 1):
		url_gen = f_url + str(i)
		print('\nСсылка преобразована:'+ Style.BRIGHT, Fore.YELLOW + url_gen + Style.RESET_ALL)
		logic_hub(url_gen, i, 1)

		print(Style.BRIGHT, Fore.CYAN + '\nСТРАНИЦА ОБРАБОТАНА' + Style.RESET_ALL)

	print(Style.BRIGHT + Fore.CYAN + '\nПРОГРАММА ЗАВЕРШИЛА РАБОТУ' + Style.RESET_ALL)

if __name__ == '__main__':
	main()
	driver.quit()